import base64
from copy import copy
import hashlib
import os
from pathlib import Path
import subprocess
import sys

from openpyxl import Workbook, load_workbook
import pytest
from sqlalchemy import inspect

from config.paths import require_test_root
from packages.backup import BackupRetentionPolicy, BackupService, RestoreVerifier
from packages.excel.template_adapter import (
    TemplateCellUpdate,
    TemplateImagePlacement,
    TemplatePreservingExporter,
)


PDF = b"%PDF-1.7\nbackup test\n%%EOF"
PNG_1X1 = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="
)


def test_backup_manifest_publish_and_restore_detect_missing_and_tampered(managed_file_env):
    env = managed_file_env
    attachment = env["service"].upload_attachment(
        product_id=env["product_id"], original_filename="drawing.pdf",
        declared_mime="application/pdf", content=PDF, actor="backup-user",
        attachment_category="DRAWING",
    )
    metadata = env["root"] / "metadata.json"
    metadata.write_text('{"schema":"0004_managed_files"}', encoding="utf-8")
    backup_service = BackupService(backup_root=env["root"] / "backups", storage=env["storage"])
    bundle = backup_service.create(
        managed_files=env["repository"].ready_files(),
        application_version="0.1.0",
        schema_revision="0004_managed_files",
        metadata_export=metadata,
    )
    assert not bundle.name.startswith(".")
    assert not tuple((env["root"] / "backups").glob("*.staging"))
    verified = RestoreVerifier().verify(bundle)
    assert verified.valid and verified.checked_files == 1

    manifest = bundle / "manifest.json"
    original_manifest = manifest.read_bytes()
    manifest.write_bytes(original_manifest.replace(b'"application_version": "0.1.0"', b'"application_version": "9.9.9"'))
    manifest_tampered = RestoreVerifier().verify(bundle)
    assert manifest_tampered.issues == ("MANIFEST_CHECKSUM_MISMATCH",)
    manifest.write_bytes(original_manifest)

    stored_copy = bundle / "files" / Path(*attachment.managed_file.storage_key.split("/"))
    stored_copy.write_bytes(b"tampered")
    tampered = RestoreVerifier().verify(bundle)
    assert not tampered.valid and any(issue.startswith("TAMPERED:") for issue in tampered.issues)
    stored_copy.unlink()
    missing = RestoreVerifier().verify(bundle)
    assert not missing.valid and any(issue.startswith("MISSING:") for issue in missing.issues)
    assert backup_service.retention_candidates(BackupRetentionPolicy(latest_n=1)) == ()


def test_excel_template_copy_preserves_source_and_bounded_template_features(tmp_path: Path):
    assert require_test_root() in tmp_path.resolve().parents
    source = tmp_path / "canonical-reference.xlsx"
    workbook = Workbook()
    try:
        sheet = workbook.active
        sheet.title = "LIST HANG"
        sheet["A1"] = "Company"
        sheet["B1"] = "Part Name"
        sheet["C2"] = "=1+1"
        sheet.merge_cells("A3:B3")
        sheet["A3"] = "Merged"
        sheet.column_dimensions["D"].hidden = True
        sheet.column_dimensions["A"].width = 24
        sheet.row_dimensions[2].height = 30
        bold_font = copy(sheet["A1"].font)
        bold_font.bold = True
        sheet["A1"].font = bold_font
        workbook.save(source)
    finally:
        workbook.close()
    image = tmp_path / "product.png"
    image.write_bytes(PNG_1X1)
    source_hash = hashlib.sha256(source.read_bytes()).hexdigest()
    output = tmp_path / "exports" / "populated.xlsx"
    result = TemplatePreservingExporter(allowed_output_root=tmp_path).export(
        template_path=source,
        output_path=output,
        cell_updates=(TemplateCellUpdate("LIST HANG", "A2", "HMS"),),
        images=(TemplateImagePlacement("LIST HANG", "B2", image, width=48, height=48, row_height=42),),
    )
    assert result.source_unchanged
    assert hashlib.sha256(source.read_bytes()).hexdigest() == source_hash
    source_book = load_workbook(source, data_only=False)
    output_book = load_workbook(output, data_only=False)
    try:
        assert source_book["LIST HANG"]["A2"].value is None
        sheet = output_book["LIST HANG"]
        assert sheet["A2"].value == "HMS"
        assert sheet["C2"].value == "=1+1"
        assert "A3:B3" in {str(cell_range) for cell_range in sheet.merged_cells.ranges}
        assert sheet.column_dimensions["D"].hidden
        assert sheet.column_dimensions["A"].width == 24
        assert sheet["A1"].font.bold
        assert len(sheet._images) == 1
        assert sheet.row_dimensions[2].height == 42
    finally:
        source_book.close()
        output_book.close()
    with pytest.raises(ValueError, match="artifact root"):
        TemplatePreservingExporter(allowed_output_root=tmp_path).export(
            template_path=source,
            output_path=Path.cwd() / "forbidden-output.xlsx",
        )
    assert not (Path.cwd() / "forbidden-output.xlsx").exists()


def test_fresh_and_upgrade_path_alembic_include_stage6_tables(tmp_path: Path):
    for name, start in (("fresh.sqlite", None), ("upgrade.sqlite", "0003_qc_packing_delivery")):
        database = tmp_path / name
        if start:
            _run_clean_alembic(database, "upgrade", start)
        _run_clean_alembic(database, "upgrade", "head")
        from sqlalchemy import create_engine
        tables = set(inspect(create_engine(f"sqlite:///{database.as_posix()}")).get_table_names())
        assert {"managed_files", "product_file_relations"} <= tables


def _run_clean_alembic(database: Path, operation: str, revision: str) -> None:
    code = (
        "import sys; "
        "from alembic.config import Config; "
        "from alembic import command; "
        "cfg=Config('alembic.ini'); "
        "cfg.set_main_option('sqlalchemy.url', 'sqlite:///' + sys.argv[1]); "
        f"command.{operation}(cfg, sys.argv[2])"
    )
    environment = os.environ.copy()
    environment["PYTHONDONTWRITEBYTECODE"] = "1"
    subprocess.run(
        [sys.executable, "-B", "-c", code, database.as_posix(), revision],
        cwd=Path.cwd(),
        env=environment,
        check=True,
    )
