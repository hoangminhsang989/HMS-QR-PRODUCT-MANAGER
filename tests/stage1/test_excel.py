from pathlib import Path

from openpyxl import Workbook, load_workbook

from packages.application.product_service import ProductService
from packages.excel.product_excel import ProductExcelExporter, ProductExcelImporter
from packages.persistence.sqlite_product_repository import SQLiteProductRepository


def test_excel_preview_requires_explicit_confirm_and_exports_valid_workbook(tmp_path: Path):
    source = tmp_path / "input.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Mã sản phẩm", "Company", "Part Name", "Quantity", "Unit", "Status"])
    sheet.append(["SP-2026-000010", "HMS", "Plate", 10, "pcs", "NEW"])
    sheet.append(["SP-2026-000011", "HMS", "Broken", 0, "pcs", "NEW"])
    workbook.save(source)
    preview = ProductExcelImporter().preview(source)
    assert preview.valid_rows == 1 and preview.invalid_rows == 1 and not preview.can_confirm
    service = ProductService(SQLiteProductRepository(tmp_path / "db.sqlite"))
    clean = tmp_path / "clean.xlsx"
    wb = Workbook(); ws = wb.active
    ws.append(["Product Code", "Company", "Part Name", "Quantity", "Unit", "Status"])
    ws.append(["SP-2026-000010", "HMS", "Plate", 10, "pcs", "NEW"])
    wb.save(clean)
    accepted = ProductExcelImporter().preview(clean)
    imported = ProductExcelImporter().confirm(accepted, service, actor="excel-user")
    assert len(imported) == 1
    output = ProductExcelExporter().export(imported, tmp_path / "export.xlsx")
    assert output.exists()
    check = load_workbook(output, read_only=True)
    try:
        assert check.active.max_row == 2
    finally:
        check.close()
