"""Generic Product Master workbook import preview and export."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook

from packages.application.product_service import ProductService
from packages.domain.product import Product, ProductStatus, ProductValidationError
from packages.domain.repository import DuplicateProductCode


HEADERS = ["Product Code", "Company", "Part Name", "Quantity", "Unit", "Material", "Requester",
           "Surface Treatment", "Outsourced", "Size", "Notes", "Delivery Schedule", "Status"]
HEADER_ALIASES = {
    "productcode": "product_code", "product code": "product_code", "mã sản phẩm": "product_code",
    "company": "company", "customer": "company", "công ty": "company", "khách hàng": "company",
    "partname": "part_name", "part name": "part_name", "tên chi tiết": "part_name",
    "quantity": "quantity", "qty": "quantity", "số lượng": "quantity",
    "unit": "unit", "đơn vị": "unit", "material": "material", "vật liệu": "material",
    "requester": "requester", "người đặt": "requester", "surfacetreatment": "surface_treatment",
    "surface treatment": "surface_treatment", "xử lý bề mặt": "surface_treatment",
    "outsourced": "outsourced", "gia công ngoài": "outsourced", "size": "size", "kích thước": "size",
    "notes": "notes", "ghi chú": "notes", "delivery schedule": "delivery_schedule",
    "lịch giao hàng": "delivery_schedule", "status": "status", "trạng thái": "status",
}


@dataclass(frozen=True, slots=True)
class ImportRowResult:
    row_number: int
    values: dict[str, object]
    errors: tuple[str, ...] = ()
    warnings: tuple[str, ...] = ()
    duplicate: bool = False


@dataclass(frozen=True, slots=True)
class ImportPreview:
    rows: tuple[ImportRowResult, ...]
    valid_rows: int
    invalid_rows: int
    warnings: int
    duplicates: int

    @property
    def can_confirm(self) -> bool:
        return self.valid_rows > 0 and self.invalid_rows == 0 and self.duplicates == 0


class ProductExcelImporter:
    def preview(self, path: str | Path, *, existing_codes: Iterable[str] = ()) -> ImportPreview:
        workbook = load_workbook(path, read_only=True, data_only=False)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return ImportPreview((), 0, 0, 0, 0)
        columns = self._map_headers(rows[0])
        known = {str(code).strip().upper() for code in existing_codes}
        results: list[ImportRowResult] = []
        for row_number, raw in enumerate(rows[1:], start=2):
            if all(value in (None, "") for value in raw):
                continue
            values: dict[str, object] = {}
            errors: list[str] = []
            for index, field in columns.items():
                if field:
                    values[field] = raw[index] if index < len(raw) else None
            values = self._convert(values, errors)
            code = str(values.get("product_code") or "").strip().upper()
            duplicate = bool(code and code in known)
            if duplicate:
                errors.append("Mã sản phẩm đã tồn tại; không tự động ghi đè.")
            if code:
                known.add(code)
            results.append(ImportRowResult(row_number, values, tuple(errors), (), duplicate))
        valid = sum(not result.errors for result in results)
        return ImportPreview(tuple(results), valid, len(results) - valid, sum(len(r.warnings) for r in results), sum(r.duplicate for r in results))

    def confirm(self, preview: ImportPreview, service: ProductService, *, actor: str) -> tuple[Product, ...]:
        if not preview.can_confirm:
            raise ValueError("Không thể import khi preview còn lỗi hoặc trùng mã.")
        created: list[Product] = []
        for row in preview.rows:
            created.append(service.create_product(actor=actor, **row.values))
        return tuple(created)

    @staticmethod
    def _map_headers(header: tuple[object, ...]) -> dict[int, str | None]:
        mapping: dict[int, str | None] = {}
        for index, value in enumerate(header):
            key = "".join(str(value or "").strip().lower().split())
            mapping[index] = HEADER_ALIASES.get(key) or HEADER_ALIASES.get(str(value or "").strip().lower())
        return mapping

    @staticmethod
    def _convert(values: dict[str, object], errors: list[str]) -> dict[str, object]:
        if "quantity" in values:
            try:
                values["quantity"] = Decimal(str(values["quantity"]).replace(",", ""))
                if values["quantity"] <= 0:
                    raise InvalidOperation
            except (InvalidOperation, TypeError, ValueError):
                errors.append("Quantity phải là số lớn hơn 0.")
        if "outsourced" in values:
            values["outsourced"] = str(values["outsourced"]).strip().lower() in {"1", "true", "yes", "y", "có", "x"}
        if isinstance(values.get("delivery_schedule"), datetime):
            values["delivery_schedule"] = values["delivery_schedule"].date()
        elif values.get("delivery_schedule"):
            try:
                values["delivery_schedule"] = date.fromisoformat(str(values["delivery_schedule"]).strip())
            except ValueError:
                errors.append("Ngày giao phải theo YYYY-MM-DD.")
        try:
            if values.get("status"):
                values["status"] = ProductStatus(str(values["status"]).strip().upper())
        except ValueError:
            errors.append("Trạng thái không hợp lệ.")
        return values


class ProductExcelExporter:
    def export(self, products: Iterable[Product], path: str | Path) -> Path:
        target = Path(path)
        target.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Products"
        sheet.append(HEADERS)
        for product in products:
            sheet.append([product.product_code, product.company, product.part_name, float(product.quantity), product.unit,
                          product.material, product.requester, product.surface_treatment, product.outsourced, product.size,
                          product.notes, product.delivery_schedule, product.status.value])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        widths = [18, 22, 26, 12, 12, 18, 18, 24, 12, 18, 30, 18, 16]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + index)].width = width
        workbook.save(target)
        return target
