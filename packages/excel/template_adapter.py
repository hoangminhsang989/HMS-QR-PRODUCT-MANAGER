"""Template-preserving Excel export foundation.

This adapter deliberately separates template export from generic import.  R008
does not claim exact business-workbook fidelity until the canonical workbook is
provided and tested.
"""

from __future__ import annotations

from dataclasses import dataclass
import hashlib
import os
from pathlib import Path
import shutil
from typing import Iterable
from uuid import uuid4

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage

from config.paths import TEST_ROOT


@dataclass(frozen=True, slots=True)
class TemplateCellUpdate:
    sheet_name: str
    cell: str
    value: object


@dataclass(frozen=True, slots=True)
class TemplateImagePlacement:
    sheet_name: str
    anchor: str
    image_path: Path
    width: int | None = None
    height: int | None = None
    row_height: float | None = None


@dataclass(frozen=True, slots=True)
class TemplateExportResult:
    output_path: Path
    source_sha256_before: str
    source_sha256_after: str
    source_unchanged: bool


class TemplatePreservingExporter:
    def __init__(self, *, allowed_output_root: str | Path = TEST_ROOT) -> None:
        self.allowed_output_root = Path(allowed_output_root).resolve(strict=False)
        self.allowed_output_root.mkdir(parents=True, exist_ok=True)

    def export(
        self,
        *,
        template_path: str | Path,
        output_path: str | Path,
        cell_updates: Iterable[TemplateCellUpdate] = (),
        images: Iterable[TemplateImagePlacement] = (),
    ) -> TemplateExportResult:
        source = Path(template_path).resolve(strict=True)
        output = Path(output_path).resolve(strict=False)
        if not source.is_file() or source.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise ValueError("Reference template must be an existing XLSX/XLSM file.")
        if output == source:
            raise ValueError("Canonical template cannot be modified in place.")
        if self.allowed_output_root != output and self.allowed_output_root not in output.parents:
            raise ValueError("Excel output must remain beneath the configured artifact root.")
        if output.suffix.lower() != source.suffix.lower():
            raise ValueError("Output format must match the reference template.")
        before = _sha256_file(source)
        output.parent.mkdir(parents=True, exist_ok=True)
        temporary = output.parent / f".{output.stem}.{uuid4().hex}.tmp{output.suffix}"
        shutil.copy2(source, temporary)
        updates = tuple(cell_updates)
        placements = tuple(images)
        try:
            if updates or placements:
                workbook = load_workbook(
                    temporary,
                    keep_vba=source.suffix.lower() == ".xlsm",
                    data_only=False,
                )
                try:
                    for update in updates:
                        if update.sheet_name not in workbook.sheetnames:
                            raise ValueError(f"Unknown template sheet: {update.sheet_name}")
                        workbook[update.sheet_name][update.cell] = update.value
                    for placement in placements:
                        if placement.sheet_name not in workbook.sheetnames:
                            raise ValueError(f"Unknown template sheet: {placement.sheet_name}")
                        image_path = Path(placement.image_path).resolve(strict=True)
                        image = OpenpyxlImage(image_path)
                        if placement.width is not None:
                            image.width = placement.width
                        if placement.height is not None:
                            image.height = placement.height
                        sheet = workbook[placement.sheet_name]
                        sheet.add_image(image, placement.anchor)
                        if placement.row_height is not None:
                            row = sheet[placement.anchor].row
                            sheet.row_dimensions[row].height = placement.row_height
                    workbook.save(temporary)
                finally:
                    workbook.close()
            os.replace(temporary, output)
        finally:
            temporary.unlink(missing_ok=True)
        after = _sha256_file(source)
        if before != after:
            raise RuntimeError("Reference template changed during export.")
        return TemplateExportResult(output, before, after, True)


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
